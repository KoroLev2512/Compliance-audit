import copy
import os
from models.task import *
from models.theme import *
from models.worker import *
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from params_calc import calculate_params
from gant_diag import build_gant
import pandas as pd
from src.visualization_graph import generate_svg_graph


class ObjectWrapper:
    value = None


def get_tasks_batch(tasks, workers_amount):
    # находим медиану от времени выполнения темы
    tasks_sum_median = sum(t.duration for t in tasks) / workers_amount
    # создаем партиции, в которые будем добавлять задачи
    partitions = [[] * workers_amount for _ in range(workers_amount)]

    # заполняем каждую партицию, пока сумма времени выполнения задач в ней меньше медианы
    for partition in partitions:
        while sum(t.duration for t in partition) < tasks_sum_median and len(tasks) > 0:
            partition.append(tasks[0])
            tasks.pop(0)

    return partitions


def assign_workers_in_theme(theme):
    # если в теме нет работников, то ничего не делаем
    if len(theme.workers) == 0:
        return

    # проверяем заняты ли все воркеры
    for worker in theme.workers:
        # если у воркера есть задача, то мы смотрим на следующего
        if worker.current_task:
            continue
        else:
            # если у кого-то нет задачи, то идем назначать задачу
            break
    else:
        # у всех воркеров есть задачи, значит ничего не делаем
        return

    # если больше одного работника, то мы сортируем воркеров по коэфу оперативности по убыванию
    if len(theme.workers) > 1:
        theme.workers.sort(key=lambda w: w.themes_efficiency[theme.name - 1], reverse=True)

    # получаем батчи жадным алгоритмом для каждого воркера
    batches = get_tasks_batch(theme.tasks.copy(), len(theme.workers))
    # список воркеров, которые покинут тему, если не получат задачу
    no_task_workers = []

    # идем по воркерам для назначения задач
    for i in range(len(theme.workers)):
        current_worker = theme.workers[i]
        # если у текущего воркера нет задачи
        if current_worker.current_task is None:
            # если на текущего воркера нет задачи из распределения
            if len(batches[i]) == 0:
                # находим следущую тему для воркера
                new_theme_index = find_next_theme_index(current_worker)

                # если тема не финальная
                if new_theme_index is not None:
                    # добавляем воркера в список свободных
                    no_task_workers.append(current_worker)
                    # назначаем задачи в теме
                    assign_workers_in_theme(themes_copy[new_theme_index])
            else:
                # если для текущего работника есть задача в батче, то мы ее назначем на него. удаляем задачу из темы
                current_worker.current_task = batches[i][0]
                current_worker.history.append(copy.deepcopy(current_worker.current_task))
                theme.tasks.remove(batches[i][0])

    # убираем из темы работников, которые ушли на другую тему
    for free_worker in no_task_workers:
        theme.remove_worker(free_worker)

    return


def find_next_theme_index(worker):
    # находим самую эффективную тему для воркера
    max_theme_index, max_efficiency = max(enumerate(worker.themes_efficiency), key=lambda pair: pair[1])

    '''так как мы ищем следующую тему + в данной теме нет свободных задач, 
    то мы воркеру зануляем оперативность по данной теме'''
    worker.themes_efficiency[max_theme_index] = 0

    while max_efficiency != 0:
        # находим следующую самую эффективную тему для воркера
        max_theme_index, max_efficiency = max(enumerate(worker.themes_efficiency), key=lambda pair: pair[1])

        # если максимальная оперативность равна 0, то он сделал все задачи из тем 1-9
        if max_efficiency == 0:
            return None

        # если в новой теме есть задачи, то мы назначаем воркера
        if len(themes_copy[max_theme_index].tasks) > 0:
            themes_copy[max_theme_index].add_worker(worker)
            return max_theme_index
        else:
            # если в новой теме нет задач, то мы зануляем оперативность и ищем следующую тему
            worker.themes_efficiency[max_theme_index] = 0


def input_data(bd1_file_path: ObjectWrapper, workers: ObjectWrapper, themes: ObjectWrapper):
    bd1_file_path.value = filedialog.askopenfilename(title="Выберите файл БД 1",
                                                     filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if bd1_file_path.value == '':
        messagebox.showwarning('Внимание', 'Вы не выбрали файл БД 1')
        return
    # Считываем задачи
    tasks_df = pd.read_excel(bd1_file_path.value, dtype={'№ работы': str})
    tasks_df = tasks_df[["№ работы", "Длительность работы, мин"]]

    themes.value = [Theme(None, None)] * max([int(e[1]['№ работы'].split('.')[0]) for e in tasks_df.iterrows()])

    for index, row in tasks_df.iterrows():
        theme_name = int(row['№ работы'].split('.')[0])
        task = Task(row['№ работы'], row['Длительность работы, мин'], theme_name)
        if themes.value[theme_name - 1].name is None:
            themes.value[theme_name - 1] = Theme(theme_name, [task])
        else:
            themes.value[theme_name - 1].tasks.append(task)

    for theme in themes.value:
        theme.tasks.sort(key=lambda x: x.duration, reverse=True)

    messagebox.showinfo("Успешно", "Данные из БД 1 успешно считаны")

    bd2_file_path = filedialog.askopenfilename(title="Выберите файл БД 2",
                                               filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if bd2_file_path == '':
        messagebox.showwarning('Внимание', 'Вы не выбрали файл БД 2')
        return

    # считываем работников
    workbook = load_workbook(bd2_file_path)
    num_sheets = len(workbook.worksheets)

    workers_df = pd.read_excel(bd2_file_path, sheet_name=num_sheets - 1, dtype={'Эксперт': str})

    workers.value = []
    for index, row in workers_df.iterrows():
        worker = Worker(row['Эксперт'], [], None, [float(e) for e in workers_df.iloc[index].to_numpy()[1:]])
        if worker.themes_efficiency.count(1) > 0:
            worker.is_president = True
        workers.value.append(worker)

    messagebox.showinfo("Успешно", "Данные из БД 2 успешно считаны")
    # если функция была вызвана без инициализации интерфейса
    try:
        interface.geometry('500x100')
        # Изменение текста кнопки
        input_data_button.config(text="Данные считаны.\nХотите выбрать БД файлы повторно?")
    except NameError:
        pass


def optimize(bd1_file_path: ObjectWrapper, workers: ObjectWrapper, themes: ObjectWrapper):
    global workers_copy, themes_copy

    optimisation_df, busy_workers_every_hour, output_file_path = optimize_df(workers.value, themes.value)

    if output_file_path is None:
        return

    # Загрузка данных из Excel
    bd1_df = pd.read_excel(bd1_file_path.value,
                           dtype={'№ работы': str, 'Предшествующая работа': str, 'Следующая работа': str})

    params_bd1_df, critical_path_nodes_bd1, critical_path_length_bd1 = calculate_params(bd1_df)
    params_optimisation_df, critical_path_nodes_optimisation, critical_path_length_optimisation = calculate_params(
        optimisation_df)

    max_crit_len = max(len(critical_path_nodes_bd1), len(critical_path_nodes_optimisation))

    crit_path_df = pd.DataFrame({
        'Критический путь до оптимизации, время': [critical_path_length_bd1] + [None] * (max_crit_len - 1),
        'Критический путь до оптимизации': critical_path_nodes_bd1 + [None] * (
                max_crit_len - len(critical_path_nodes_bd1)),
        'Критический путь после оптимизации, время': [critical_path_length_optimisation] + [None] * (max_crit_len - 1),
        'Критический путь после оптимизации': critical_path_nodes_optimisation + [None] * (
                max_crit_len - len(critical_path_nodes_optimisation))
    })

    # датафрейм сколько работников занято в конкретный час
    busy_workers_every_hour_df = pd.DataFrame(list(busy_workers_every_hour.items()),
                                              columns=["Который час", "Сколько воркеров занято"])

    # Создание словаря для хранения данных
    worker_busy_in_topic = {}

    for worker in workers_copy:
        theme_time = {}
        for task in worker.history:
            theme_key = f'Тема {task.theme}'
            if theme_key in theme_time:
                theme_time[theme_key] += task.duration
            else:
                theme_time[theme_key] = task.duration
        if worker.is_president:
            worker_busy_in_topic[f'Эксперт {worker.name} (председатель)'] = theme_time
        else:
            worker_busy_in_topic[f'Эксперт {worker.name}'] = theme_time

    # Создание датафрейма из собранных данных
    worker_busy_in_topic_df = pd.DataFrame(worker_busy_in_topic).T.fillna(0)

    # Функция для извлечения числовой части из названия темы
    def extract_number(theme_name):
        return int(theme_name.split()[1])

    # Сортировка столбцов по числовой части их названий
    sorted_columns = sorted(worker_busy_in_topic_df.columns, key=extract_number)
    worker_busy_in_topic_df = worker_busy_in_topic_df.reindex(sorted_columns, axis=1)

    # Записываем DataFrame в Excel с использованием openpyxl
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        optimisation_df.to_excel(writer, sheet_name='Оптимизация', index=False)  # Первая страница
        params_bd1_df.to_excel(writer, sheet_name='До оптимизации', index=False)  # Вторая страница
        params_optimisation_df.to_excel(writer, sheet_name='После оптимизации', index=False)  # Третья страница
        crit_path_df.to_excel(writer, sheet_name='Критические пути', index=False)  # Четвертая страница
        busy_workers_every_hour_df.to_excel(writer, sheet_name='Статистика', index=False)  # Пятая страница
        worker_busy_in_topic_df.to_excel(writer, sheet_name='Время воркеров по темам')  # Шестая страница

    # Открытие сохраненного файла
    wb = load_workbook(output_file_path)

    # Применение выравнивания и автоширины ко всем листам
    alignment = Alignment(horizontal='center', vertical='center')
    for ws in wb.worksheets:  # Проходим по всем листам в рабочей книге
        # Применение выравнивания ко всем ячейкам
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        # Расширение колонок для автоматического вмещания текста
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    print(e)
                    pass
            adjusted_width = max_length + 2  # Добавляем небольшой отступ
            ws.column_dimensions[column_letter].width = adjusted_width

    # Сохранение изменений
    wb.save(output_file_path)

    output_folder_path = os.path.dirname(os.path.realpath(output_file_path))

    build_gant(bd1_df, 'Диаграмма Ганта с критическим путем до оптимизации', output_folder_path)
    build_gant(optimisation_df, 'Диаграмма Ганта с критическим путем после оптимизации', output_folder_path)

    generate_svg_graph(optimisation_df, output_folder_path)

    messagebox.showinfo("Успешно", f'Файл с оптимизацией сохранен по пути {output_file_path}')


def optimize_df(workers: ObjectWrapper, themes: ObjectWrapper, write_output_file: bool = True):
    global workers_copy, themes_copy

    # делаем копии, чтобы работать с ними. избавляемся от повторного указания БД файлов
    try:
        workers_copy = copy.deepcopy(workers)
        themes_copy = copy.deepcopy(themes)
    except NameError:
        messagebox.showerror('Ошибка', 'Вы не считали файлы БД')
        return None, None, None
    except AttributeError:
        messagebox.showerror('Ошибка', 'Вы не считали файлы БД')
        return None, None, None

    # назначаем воркеров на темы
    for worker in workers_copy:
        max_theme_index, max_efficiency = max(enumerate(worker.themes_efficiency), key=lambda pair: pair[1])
        themes_copy[max_theme_index].workers.append(worker)

    # назначаем задачи воркерам
    for theme in themes_copy:
        assign_workers_in_theme(theme)

    # начинаем симуляцию
    current_time = 0
    # множество закончивших работу воркеров
    free_workers = set()

    # статистика сколько воркеров занято каждый час
    busy_workers_every_hour = {}

    # пока в последней теме есть задачи
    while len(free_workers) < len(workers_copy):
        if current_time % 60 == 0:
            busy_workers_every_hour[current_time // 60] = sum(1 for w in workers_copy if w.current_task)

        for worker in workers_copy:
            # если у воркера нет текущей задачи, значит он закончил все работы из тем 1-9
            if worker.current_task is None:
                # если он уже сделал 10 тему, то пропускаем его
                if worker in free_workers:
                    continue
                else:
                    # назначаем задачу из 10 темы
                    # если председатель, то 10.1
                    if worker.is_president:
                        worker.current_task = themes_copy[-1].tasks[0]
                        worker.history.append(copy.deepcopy(worker.current_task))
                        themes_copy[-1].tasks.pop(0)
                    else:
                        # если обычный воркер, то любая задача, кроме 10.1
                        worker.current_task = themes_copy[-1].tasks[-1]
                        worker.history.append(copy.deepcopy(worker.current_task))
                        themes_copy[-1].tasks.pop()

            # уменьшаем длительность задачи на 1 минуту
            worker.current_task.duration -= 1

            # если задача выполнена
            if worker.current_task.duration == 0:
                print(f'Исполнитель {worker.name} сделал задачу {worker.current_task.name}')
                # находим тему задачи
                worker_theme = worker.current_task.theme

                # снимаем задачу с пользователя
                worker.current_task = None

                # если тема выполненной задачи 10, то работник освободился
                if worker_theme == len(themes_copy):
                    free_workers.add(worker)
                    continue

                # проверяем, что в теме не осталось свободных задач
                if len(themes_copy[worker_theme - 1].tasks) == 0:
                    # удаляем из темы воркера
                    themes_copy[worker_theme - 1].remove_worker(worker)
                    # ищем следующую подходящую воркеру тему по его коэфу оперативности
                    find_next_theme_index(worker)

        # идем по темам и распределяем задачи
        for theme in themes_copy:
            assign_workers_in_theme(theme)

        # симулируем время. + 1 минута
        current_time += 1

    print(f'Симуляция завершена. Общее время: {current_time}. В днях: {current_time / 480}')
    if write_output_file:
        messagebox.showinfo("Успешно",
                            f'Симуляция завершена. Общее время: {current_time}. В днях: {current_time / 480}')

        # Открытие диалогового окна для выбора места сохранения файла оптимизации
        output_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",  # Расширение по умолчанию
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],  # Типы файлов
            title="Выберите место для сохранения файла оптимизации"  # Заголовок окна
        )
    else:
        output_file_path = None
    output_data = {
        '№ работы': [],
        'Длительность работы, мин': [],
        'Предшествующая работа': [],
        'Следующая работа': [],
        'Исполнитель': []
    }

    for worker in workers_copy:
        for task_index in range(len(worker.history)):
            output_data['№ работы'].append(worker.history[task_index].name)
            output_data['Длительность работы, мин'].append(worker.history[task_index].duration)
            if task_index == 0:
                output_data['Предшествующая работа'].append('0')
            else:
                output_data['Предшествующая работа'].append(worker.history[task_index - 1].name)

            if task_index == len(worker.history) - 1:
                output_data['Следующая работа'].append(str(len(themes_copy) + 1))
            else:
                output_data['Следующая работа'].append(worker.history[task_index + 1].name)

            output_data['Исполнитель'].append(worker.name)

    optimisation_df = pd.DataFrame(output_data)
    # Добавляем временную колонку с преобразованными значениями
    optimisation_df['parsed'] = optimisation_df['№ работы'].apply(lambda a: list(map(int, a.split('.'))))

    # Сортируем DataFrame по временной колонке
    optimisation_df = optimisation_df.sort_values(by='parsed')

    # Удаляем временную колонку
    optimisation_df = optimisation_df.drop(columns=['parsed'])

    optimisation_df.insert(loc=0, column='№ п/п', value=list(range(1, sum(len(e.history) for e in workers_copy) + 1)))

    return optimisation_df, busy_workers_every_hour, output_file_path


if __name__ == "__main__":
    interface = Tk()
    interface.geometry('300x70')
    interface.title('Compliance Audit v1')

    filepath_obj = ObjectWrapper()
    workers_obj = ObjectWrapper()
    themes_obj = ObjectWrapper()
    # Создание стиля для кнопок
    style = ttk.Style()
    style.configure('TButton',
                    font=('Helvetica', 12),  # Изменяем шрифт и размер
                    padding=10,  # Устанавливаем отступы
                    relief='flat',  # Убираем границу для плоского стиля
                    background='#4CAF50',  # Цвет фона кнопки
                    foreground='black',  # Цвет текста
                    borderwidth=1)  # Ширина границы

    # Создание кнопок

    input_data_button = ttk.Button(interface, text="Ввод данных",
                                   command=lambda: input_data(filepath_obj, workers_obj, themes_obj), style='TButton')
    input_data_button.grid(column=0, row=0, padx=10, pady=10)

    optimization_button = ttk.Button(interface, text="Оптимизация",
                                     command=lambda: optimize(filepath_obj, workers_obj, themes_obj), style='TButton')
    optimization_button.grid(column=1, row=0, padx=10, pady=10)

    interface.mainloop()
