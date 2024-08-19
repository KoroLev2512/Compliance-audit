import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import random
import logging

logging.basicConfig(level=logging.INFO)


def generate_work_numbers(num_topics, max_mandatory_subtasks, max_optional_subtasks, optional_task_probability=0.1):
    work_numbers = []
    for topic in range(1, num_topics + 1):  # Учитываем все темы, включая последнюю
        num_mandatory_subtasks = random.randint(1, max_mandatory_subtasks)
        for mandatory_subtask in range(1, num_mandatory_subtasks + 1):
            work_number = f"{topic}.{mandatory_subtask}"
            work_numbers.append(work_number)

            # Проверяем, если текущая тема не последняя, тогда генерируем опциональные подзадачи
            if topic < num_topics and random.random() < optional_task_probability:
                num_optional_subtasks = random.randint(1, max_optional_subtasks)
                for optional_subtask in range(1, num_optional_subtasks + 1):
                    work_number_optional = f"{work_number}.{optional_subtask}"
                    work_numbers.append(work_number_optional)

    work_numbers.sort(key=lambda x: list(map(int, x.split('.'))))
    return work_numbers


def generate_duration():
    if random.random() < 0.05:
        return random.randint(100, 1500)
    else:
        return random.choice(range(5, 101, 5))


def save_dataframe(df, file_path):
    try:
        if file_path.endswith('.xlsx'):
            df.to_excel(file_path, sheet_name='Лист1', index=False)
            wb = load_workbook(file_path)
            ws = wb['Лист1']
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            wb.save(file_path)
        elif file_path.endswith('.csv'):
            df.to_csv(file_path, index=False)
        else:
            logging.error("Неподдерживаемый формат файла. Используйте .xlsx или .csv.")
    except Exception as e:
        logging.error(f"Ошибка при сохранении файла: {e}")


def create_excel_random(num_topics, max_mandatory_subtasks, max_optional_subtasks, expert_ids):
    tasks = []
    while len(tasks) < 100:  # Условие, чтобы задач было минимум 100
        tasks = generate_work_numbers(num_topics, max_mandatory_subtasks, max_optional_subtasks)

    # Создание задач для последней темы
    if num_topics > 0:
        last_topic_tasks = [f"{num_topics}.{i + 1}" for i in range(len(expert_ids))]
        tasks = [task for task in tasks if int(task.split('.')[0]) < num_topics] + last_topic_tasks

    # Создание столбцов "Предшествующая работа" и "Следующая работа"
    previous_work = ['0'] + tasks[:-1]
    next_work = tasks[1:] + [f"{num_topics + 1}"]

    # Проверка длины списка задач
    num_tasks = len(tasks)
    durations = [generate_duration() for _ in range(num_tasks)]

    # Проверка длины всех списков
    assert len(durations) == num_tasks, "Длину списка 'durations' не совпадает с числом задач"
    assert len(previous_work) == num_tasks, "Длину списка 'previous_work' не совпадает с числом задач"
    assert len(next_work) == num_tasks, "Длину списка 'next_work' не совпадает с числом задач"

    # Выбор уникальных экспертов для задач последней темы
    executors = [random.choice(expert_ids) for _ in tasks]
    if num_topics > 0:
        last_topic_experts = random.sample(expert_ids, len(last_topic_tasks))
        for i, task in enumerate(tasks):
            if int(task.split('.')[0]) == num_topics:
                executors[i] = last_topic_experts.pop(0)

    assert len(executors) == num_tasks, "Длину списка 'executors' не совпадает с числом задач"

    df = pd.DataFrame({
        '№ п/п': range(1, len(tasks) + 1),
        '№ работы': tasks,
        'Длительность работы, мин': durations,
        'Предшествующая работа': previous_work,
        'Следующая работа': next_work,
        'Исполнитель': executors
    })

    save_dataframe(df, 'БД1_тест.xlsx')
    save_dataframe(df, 'БД1_тест.csv')


def get_input(prompt, min_value, max_value):
    while True:
        try:
            value = int(input(prompt))
            if min_value <= value <= max_value:
                return value
            else:
                print(f"Введите число от {min_value} до {max_value}.")
        except ValueError:
            print("Пожалуйста, введите целое число.")


def main():
    num_experts = get_input("Введите количество экспертов (от 2 до 20): ", 2, 20)
    num_topics = get_input("Введите количество тем (от 2 до 20): ", 2, 20)

    expert_ids = random.sample(range(1, 101), num_experts)
    coefficients = [round(random.uniform(0, 1), 5) for _ in range(num_experts * (num_topics - 1))]
    coefficients_matrix = [coefficients[i * (num_topics - 1):(i + 1) * (num_topics - 1)] for i in range(num_experts)]

    selected_expert = random.randint(0, num_experts - 1)
    for i in range(num_topics - 1):
        coefficients_matrix[selected_expert][i] = 1.0 if i in (0, num_topics - 2) else 0.0

    for i in range(num_experts):
        if i != selected_expert:
            coefficients_matrix[i][0] = 0.0
            coefficients_matrix[i][num_topics - 2] = 0.0

    df_db2_sheet = pd.DataFrame(coefficients_matrix, columns=[f"Тема {i + 1}" for i in range(num_topics - 1)],
                                index=expert_ids)
    df_db2_sheet.index.name = 'Эксперт'
    df_db2_sheet.reset_index(inplace=True)

    print("\nТаблица коэффициентов:")
    print(df_db2_sheet)

    save_dataframe(df_db2_sheet, 'БД2_тест.xlsx')
    save_dataframe(df_db2_sheet, 'БД2_тест.csv')

    max_mandatory_subtasks = 6
    max_optional_subtasks = 10
    create_excel_random(num_topics, max_mandatory_subtasks, max_optional_subtasks, expert_ids)


if __name__ == "__main__":
    main()
    print("Файлы успешно созданы")
