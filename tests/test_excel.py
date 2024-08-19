import pytest
import os
import pandas as pd
from openpyxl import load_workbook

from src.excel_utils import save_dataframe, generate_db2_data, generate_db1_data


@pytest.fixture
def cleanup_files():
    yield
    for file in ['БД2_тест.xlsx', 'БД2_тест.csv', 'БД1_тест.xlsx', 'БД1_тест.csv']:
        if os.path.exists(file):
            os.remove(file)


def test_save_dataframe_xlsx(cleanup_files):
    df = pd.DataFrame(generate_db2_data())
    save_dataframe(df, 'БД2_тест.xlsx')

    assert os.path.exists('БД2_тест.xlsx')

    wb = load_workbook('БД2_тест.xlsx')
    ws = wb.active

    assert ws['A1'].alignment.horizontal == 'center'
    assert ws['A1'].alignment.vertical == 'center'


def test_save_dataframe_csv(cleanup_files):
    df = pd.DataFrame(generate_db2_data())
    save_dataframe(df, 'БД2_тест.csv')

    assert os.path.exists('БД2_тест.csv')

    df_loaded = pd.read_csv('БД2_тест.csv')
    pd.testing.assert_frame_equal(df, df_loaded)


def test_generate_db2_data():
    data = generate_db2_data()
    assert isinstance(data, dict)
    assert len(data['Эксперт']) == 6


def test_generate_db1_data():
    data = generate_db1_data()
    assert isinstance(data, dict)
    assert len(data['№ п/п']) == 266
